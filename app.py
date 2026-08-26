# archivo: app.py
from flask import Flask, request, render_template_string
from datetime import datetime
import math
import pytz
import openpyxl
import os

app = Flask(__name__)

# Coordenadas del taller (ejemplo: Villa María del Triunfo, Lima)
TALLER_LAT = -12.173043127500463
TALLER_LON = -76.91693852892678
RADIO_PERMITIDO = 0.05  # en km (50 metros)

# Definir zona horaria de Lima
tz = pytz.timezone("America/Lima")

def distancia(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def guardar_excel(nombre, dni, fecha, hora, tipo, lat, lon):
    archivo = "asistencia.xlsx"
    if not os.path.exists(archivo):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"
        ws.append(["Nombre", "DNI", "Fecha", "Hora", "Tipo", "Latitud", "Longitud"])
        wb.save(archivo)

    wb = openpyxl.load_workbook(archivo)
    ws = wb["Registros"]

    # Validar que no registre más de una vez por día en cada tipo
    for fila in ws.iter_rows(values_only=True):
        if fila[1] == dni and fila[2] == fecha and fila[4] == tipo:
            return False  # Ya existe registro

    ws.append([nombre, dni, fecha, hora, tipo, lat, lon])
    wb.save(archivo)
    return True

@app.route("/", methods=["GET", "POST"])
def asistencia():
    if request.method == "POST":
        nombre = request.form["nombre"]
        dni = request.form["dni"]
        tipo = request.form["tipo"]  # Entrada o Salida
        lat = float(request.form["lat"])
        lon = float(request.form["lon"])
        fecha = datetime.now(tz).strftime("%Y-%m-%d")
        hora = datetime.now(tz).strftime("%H:%M:%S")

        # Validar ubicación
        if distancia(lat, lon, TALLER_LAT, TALLER_LON) > RADIO_PERMITIDO:
            return "❌ No estás en el taller, registro rechazado"

        if not guardar_excel(nombre, dni, fecha, hora, tipo, lat, lon):
            return f"❌ Ya registraste {tipo} hoy"

        return f"✅ {tipo} registrada correctamente a las {hora}"

    return render_template_string('''
        <h2>Registro de Asistencia</h2>
        <form method="post" onsubmit="return enviarUbicacion();">
            Nombre: <input type="text" name="nombre" required><br>
            DNI: <input type="text" name="dni" required><br>
            Tipo: 
            <select name="tipo" required>
                <option value="Entrada">Entrada</option>
                <option value="Salida">Salida</option>
            </select><br>
            <input type="hidden" name="lat" id="lat">
            <input type="hidden" name="lon" id="lon">
            <input type="submit" value="Marcar asistencia">
        </form>
        <script>
        function enviarUbicacion() {
            if (navigator.geolocation) {
                navigator.geolocation.getCurrentPosition(function(pos) {
                    document.getElementById("lat").value = pos.coords.latitude;
                    document.getElementById("lon").value = pos.coords.longitude;
                    document.forms[0].submit();
                });
                return false;
            } else {
                alert("Tu navegador no soporta GPS");
                return false;
            }
        }
        </script>
    ''')
    
if __name__ == "__main__":
    app.run(debug=True)
from flask import send_file

@app.route("/descargar")
def descargar():
    return send_file("asistencia.xlsx", as_attachment=True)
